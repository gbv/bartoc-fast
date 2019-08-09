""" models.py """

import time # dev
import asyncio
import re

from django.db import models # https://docs.djangoproject.com/en/2.2/ref/models/fields/#django.db.models.Field
from aiohttp import ClientSession, ClientTimeout
from openpyxl import load_workbook
from SPARQLWrapper import (SPARQLWrapper, JSON)
from urllib import parse

from .utility import Database, Entry, Result, DEF_MAXSEARCHTIME, LOCAL_APP_PATH

# queries:
class Query(models.Model): # confusing, same name as schema.Query
    """ Abstract query class """
    
    querystring = models.CharField(max_length=100)
    description = models.CharField(max_length=100)
    category = models.IntegerField()
    timeout = models.IntegerField()

    class Meta:
        abstract = True

class SparqlQuery(Query):
    """ A SPARQL query """

    class Meta:
        verbose_name_plural = 'Sparql queries'
    
    sparqlendpoint = models.ForeignKey("SparqlEndpoint", on_delete=models.CASCADE)

    def __str__(self) -> str:
        return f'{self.sparqlendpoint.name} category {self.category}'

    def update(self, searchword: str) -> str:
        """ Put the searchword into the query """
        
        return self.querystring.replace("!!!SEARCHWORD!!!", searchword)

class SkosmosQuery(Query):
    """ A SPARQL query """

    class Meta:
        verbose_name_plural = 'Skosmos queries'
    
    skosmosinstance = models.ForeignKey("SkosmosInstance", on_delete=models.CASCADE)

    def __str__(self) -> str:
        return f'{self.skosmosinstance.name} category {self.category}'

# resources:
class Resource(models.Model):
    """ Abstract resource class """

    federation = models.ForeignKey("Federation", on_delete=models.CASCADE)
    name = models.CharField(max_length=100)
    url = models.URLField()
    context = models.CharField(max_length=200) #special, wildcard

    class Meta:
        abstract = True

    def __str__(self) -> str:
        return f'{self.name}'

    async def main(self,
                   session: ClientSession,
                   searchword: str,
                   category: int = 0) -> Result:
        """ Coroutine: send query to resource """

        try:
            result = await self.search(session, searchword, category)
            return result
        
        except asyncio.TimeoutError:
            print(f'FETCH {self.name} run out of time!') # dev
            return Result(self.name, None, category)

class SkosmosInstance(Resource):
    """ A Skosmos instance """

    def select(self, category: int) -> SkosmosQuery:
        """ Select Skosmos query by category """ 
        
        return SkosmosQuery.objects.get(skosmosinstance=self, category=category)

    def prepare(self, searchword: str) -> str:
        """ Prepare searchword for search """
        
        # remove wildcard as in color* and col*r, applies to Legilux:
        if "wildcard" in self.context:
            return parse.quote(searchword.replace("*", ""))
            
        return parse.quote(searchword)   

    async def search(self,
                     session: ClientSession,
                     searchword: str,
                     category: int = 0) -> Result:
        """ Coroutine: send query to Skosmos instance """

        start = time.time() # dev
        query = self.select(category)

        searchword = self.prepare(searchword)
        
        if query.timeout == 1:
            return Result(self.name, None, category)

        restapi = self.url + query.querystring + searchword
        
        async with session.get(restapi) as response:
            
            data = await response.json()

            end = time.time() # dev
            print(f'FETCH {self.name} took {end - start}') # dev

            return Result(self.name, data, category)

class SparqlEndpoint(Resource):
    """ A SPARQL endpoint """

    def select(self, category: int) -> SparqlQuery:
        """ Select SPARQL query by category """ 
        
        return SparqlQuery.objects.get(sparqlendpoint=self, category=category)

    def prepare(self, searchword: str) -> str: 
        """ Prepare searchword for search """

        # remove excess whitespace:
        " ".join(searchword.split())

        # open compounds ("gas station"):
        searchword = searchword.replace(" ", " AND ")
        
        # remove special characters (umlaute, etc), applies to virtuoso endpoints (LuSTRE):
        if "special" in self.context:
            searchword = re.sub('[^A-Za-z0-9\s]+', '', searchword)
            " ".join(searchword.split())
            
        return searchword

    async def search(self,
                     session: ClientSession,
                     searchword: str,
                     category: int = 0) -> Result:
        """ Coroutine: send query to SPARQL endpoint """

        start = time.time() # dev
        query = self.select(category)

        searchword = self.prepare(searchword)

        if query.timeout == 1:
            return Result(self.name, None, category)

        querystring = query.update(searchword)

        # use SPARQLWrapper to parse querystring:
        wrapper = SPARQLWrapper(self.url)
        wrapper.setQuery(querystring)
        wrapper.setReturnFormat(JSON)
        request = wrapper._createRequest()
        parsed_query = request.get_full_url()

        async with session.get(parsed_query) as response:

            data = await response.json()

            end = time.time() # dev
            print(f'FETCH {self.name} took {end - start}') # dev

            return Result(self.name, data, category)

# federation: 
class Federation(models.Model):
    """ The federation of resources (there is only one) """
    created = models.DateTimeField(auto_now=True)

    class Meta:
        verbose_name_plural = 'federation'

    def __str__(self) -> str:
        return 'Main federation'

    def populate_skosmosinstances(self) -> None: 
        """ Populate federation with Skosmos instances and their queries """

        SkosmosInstance.objects.all().delete()

        skosmosinstances = load_workbook(LOCAL_APP_PATH + "fixtures/skosmosinstances.xlsx") # required for unittest; simply fixtures/skosmosinstances.xlsx in production

        for ws in skosmosinstances:
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):   
                instance = SkosmosInstance(federation=self,
                                                  name=row[0],
                                                  url=row[1],
                                                  context=row[2])
                instance.save()

                query = SkosmosQuery(skosmosinstance=instance, # constructs query of cat 0
                                     querystring="/rest/v1/search?query=",
                                     description="NA",
                                     category=0,
                                     timeout=row[3])
                query.save()

    def populate_sparqlendpoints(self) -> None:
        """ Populate federation with SPARQL endpoints and their queries """

        SparqlEndpoint.objects.all().delete()

        sparqlendpoints = load_workbook(LOCAL_APP_PATH + "fixtures/sparqlendpoints.xlsx") # required for unittest; fixtures/sparqlendpoints.xlsx in production

        for ws in sparqlendpoints:

            endpoint = SparqlEndpoint(federation=self,
                                      name=ws['A3'].value,
                                      url=ws['B3'].value,
                                      context=ws['C3'].value)
            endpoint.save()
            
            for row in ws.iter_rows(min_row=3, min_col=4, max_col=7, values_only=True):
                query = SparqlQuery(sparqlendpoint=endpoint,
                                    querystring=row[0],
                                    description=row[1],
                                    category=row[2],
                                    timeout=row[3])
                query.save()

    def populate(self) -> None:
        """ Populate federation with all resources """

        self.populate_skosmosinstances()
        self.populate_sparqlendpoints()

            

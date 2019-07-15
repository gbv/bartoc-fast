""" utility.py """

from __future__ import annotations # see https://www.python.org/dev/peps/pep-0563/
from typing import List, Set, Dict, Tuple, Optional, Union

from openpyxl import load_workbook

LOCAL_APP_PATH = "//itsc-pg2.storage.p.unibas.ch/ub-home$/hinder0000/Documents/GitHub/bartocgraphql/django/bartocgraphql/" # local app path
STANDARD_TIME = 5 # standard search time in seconds
STANDARD_DUPLICATES = False # standard value for duplicates argument

class Database:
    """ A database """
    
    def __init__(self, entries: set) -> None:
        self.entries = entries

    def set_entries(self, entries: set) -> None:
        self.entries = entries

    def add_entries(self, entries: set) -> None:
        self.entries = entries.union(self.entries)

class Entry:
    """ A named and linked entry of a database """
    
    def __init__(self, name: str, url: str) -> None:
        self.name = name
        self.url = url

    def set_name(self, name: str) -> None:
        self.name = name
         
    def set_url(self, url: str) -> None:
        self.url = url

class Result:
    """ The result of a search """
    
    def __init__(self, name: str, data: Union[dict, None], category: int = 0) -> None:
        self.name = name            # of the endpoint or instance
        self.data = data            # JSON output
        self.category = category    # of the query, if any

class MappingDatabase(Database):
    """ A collection of mappings """

    def __init__(self, entries: set = {}) -> None:
        Database.__init__(self, entries)
        self.setup()

    def setup(self) -> None:
        """ Populate the database with maps """
        
        wb = load_workbook("//itsc-pg2.storage.p.unibas.ch/ub-home$/hinder0000/Documents/GitHub/bartocgraphql/django/bartocgraphql/fixtures/mappings.xlsx") # required for unittest; fixtures/mappings.xlsx in production
        for ws in wb:
            for row in ws.iter_rows(min_row=3, min_col=1, max_col=4, values_only=True):
                field = row[0]
                uri = row[1]
                equivalences = []
                n = 1 # normative SKOS uri is equivalent to itself so it must be added to equivalences
                while n < 4:
                    if row[n] == None:
                        pass
                    else:
                        equivalences.append(row[n])
                    n += 1
                self.add_entries({Mapping(field, uri, equivalences)})

    def select_by_field(self, field: str) -> Mapping: 
        """ Select mapping by (name of) field """

        for mapping in self.entries:
            if mapping.field == field:
                return mapping
            
    def select_by_uri(self, uri: str) -> Mapping:
        """ Select mapping by uri """

        for mapping in self.entries:
            if mapping.uri in mapping.equivalences:
                return mapping

class Mapping:
    """ A mapping """
    
    def __init__(self, field: str, uri: str, equivalences: List[str]) -> None:
        self.field = field
        self.uri = uri
        self.equivalences = equivalences



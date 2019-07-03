import time                                     # dev
import asyncio

from aiohttp import ClientSession
from openpyxl import load_workbook

from .utility import Database, Entry, Result    # local

class SkosmosDatabase(Database):
    """ A collection of Skosmos instances """
    
    def __init__(self, entries: set = {}) -> None:
        Database.__init__(self, entries)
        self.setup() # setup is called on initializing class

    def setup(self) -> None:
        """ Populates the database with Skosmos instances """

        wb = load_workbook("//itsc-pg2.storage.p.unibas.ch/ub-home$/hinder0000/Documents/GitHub/bartocgraphql/django/bartocgraphql/fixtures/skosmosinstances.xlsx") # required for unittest; fixtures/skosmosinstances.xlsx in production
        for ws in wb:
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):   
                name = row[0]
                url = row[1]
                timeout = row[2]
                self.add_entries({SkosmosInstance(name, url, timeout)})

class SkosmosInstance(Entry):
    """ A Skosmos instance """
    
    def __init__(self, name: str, url: str, timeout: int) -> None:
        Entry.__init__(self, name, url)
        self.timeout = timeout

    async def search(self, session: ClientSession, searchword: str, category: int = 0) -> Result:
        """ Coroutine: calls Global_methods/get_search for searchword at instance's REST API """

        start = time.time()                                 # dev
        
        if self.timeout == 1:
            return Result(self.name, None, category)
        
        restapi = self.url + "/rest/v1/search?query=" + searchword
        async with session.get(restapi) as response:
            
            data = await response.json()

            end = time.time()                               # dev
            print(f'ASYNC {self.name} took {end - start}')  # dev

            return Result(self.name, data, category)
            

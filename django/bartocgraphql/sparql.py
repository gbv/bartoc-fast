""" sparql.py """

from __future__ import annotations
import time # dev
from typing import List, Set, Dict, Tuple, Optional
import asyncio

import aiohttp
from openpyxl import load_workbook
from SPARQLWrapper import (SPARQLWrapper, JSON)

from .utility import Database, Entry, Result # local

class SparqlDatabase(Database):
    """ A collection of SPARQL endpoints """
    
    def __init__(self, entries: set = {}) -> None:
        Database.__init__(self, entries)
        self.setup() # setup is called on initializing class

    def setup(self) -> None:
        """ Populate the database with SPARQL endpoints and their queries """

        wb = load_workbook("//itsc-pg2.storage.p.unibas.ch/ub-home$/hinder0000/Documents/GitHub/bartoc-graphql/django/bartocgraphql/fixtures/sparqlendpoints.xlsx") # required for unittest; fixtures/sparqlendpoints.xlsx in production
        for ws in wb:
            queries = []
            for row in ws.iter_rows(min_row=3, min_col=3, max_col=6, values_only=True):   
                querystring = row[0]
                description = row[1]
                category = row[2]
                timeout = row[3]
                queries.append(SparqlQuery(querystring, description, category, timeout))
            self.add_entries({SparqlEndpoint(ws['A3'].value, ws['B3'].value, queries)})

    def select(self, name: str) -> SparqlEndpoint:
        """ Select SPARQL endpoint by name """

        for endpoint in self.entries:
            if endpoint.name == name:
                return endpoint

class SparqlEndpoint(Entry):
    """ A SPARQL endpoint """
    
    def __init__(self, name: str, url: str, queries: List[SparqlQuery]) -> None:
        Entry.__init__(self, name, url)   
        self.queries = queries

    async def search(self, session: aiohttp.ClientSession, searchword: str, category: int = 0) -> Result:
        """ Coroutine: Send a query (default category 0) for searchword to the SPARQL endpoint and return the results """
        start = time.time()                                 # dev
        query = self.select(category)

        if query.timeout == 1:
            return Result(self.name, None, category)
        
        query.update(searchword)

        # we just want the flat query, not the request...
        sparqlwrapper = SPARQLWrapper(self.url)
        sparqlwrapper.setQuery(query.querystring)
        sparqlwrapper.setReturnFormat(JSON)
        flatquery = sparqlwrapper.query().geturl()

        # ...because the request is done here
        async with session.get(flatquery) as response:
            data = await response.json()
            query.reset()
            end = time.time()                               # dev
            print(f'ASYNC {self.name} took {end - start}')  # dev
            return Result(self.name, data, category)

    def select(self, category: int) -> SparqlQuery:
        """ Select query by category """
        
        for query in self.queries:
            if query.category == category:
                return query
            
class SparqlQuery:
    """ A SPARQL query (to be used with a specific endpoint) """
    
    def __init__(self, querystring: str, description: str, category: int, timeout: int) -> None:
        self.querystring = querystring
        self.unmodified = querystring
        self.description = description
        self.category = category
        self.timeout = timeout

    def update(self, searchword: str) -> None:
        """ Put the searchword into the query """
        
        self.querystring = self.querystring.replace("!!!SEARCHWORD!!!", searchword)

    def reset(self) -> None:
        """ Reset to the initial query (w/o searchword) """
        
        self.querystring = self.unmodified

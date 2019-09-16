""" maintenance.py """

import asyncio 
from typing import List, Set, Dict, Union

from openpyxl import load_workbook

from .models import Federation, SkosmosInstance, SkosmosQuery, SparqlEndpoint, SparqlQuery, LobidResource, LobidQuery, LdapiEndpoint, LdapiQuery
from .schema import Helper
from .utility import LOCAL_APP_PATH

### from .models import MODELS_RESOURCES, MODELS_FEDERATION # or make function instead of constants

class Maintenance:
    """ Maintenance tools """

    @classmethod
    def selfcheck(self) -> None:
        """ Disabe slow resources """

        resources = list(SparqlEndpoint.objects.all()) + list(SkosmosInstance.objects.all()) + list(LobidResource.objects.all()) ###

        # "logic" searchword
        logic = ["AgroVoc",
                 "Bartoc",
                 "data.ub.uio.no",
                 "Finto",
                 "GettyAAT",
                 "GettyULAN",
                 "Irstea",
                 "lobid-gnd",
                 "Loterre",
                 "LuSTRE",
                 "OZCAR-Theia",
                 "UAAV",
                 "UNESCO"]
        logic_resources = []
        for resource in resources:
            if resource.name in logic:
                logic_resources.append(resource)
        logic_results = asyncio.run(Helper.fetch(logic_resources, "logic", 0, 5), debug=True)

        # "law" searchword
        law = ["51.15.194.251",
               "DemoVoc",
               "GACS",
               "GettyTGN",
               "Legilux",
               "ScoLOMFR"]
        law_resources = []
        for resource in resources:
            if resource.name in law:
                law_resources.append(resource)
        law_results = asyncio.run(Helper.fetch(law_resources, "law", 0, 5), debug=True)

        # "trumpet" searchword
        trumpet = ["MIMO"]
        trumpet_resources = []
        for resource in resources:
            if resource.name in trumpet:
                trumpet_resources.append(resource)
        trumpet_results = asyncio.run(Helper.fetch(trumpet_resources, "trumpet", 0, 5), debug=True)

         # "blog" searchword
        blog = ["HTW Chur"]
        blog_resources = []
        for resource in resources:
            if resource.name in blog:
                blog_resources.append(resource)
        blog_results = asyncio.run(Helper.fetch(blog_resources, "blog", 0, 5), debug=True)

        # "material" searchword
        material = ["FORTH"]
        material_resources = []
        for resource in resources:
            if resource.name in material:
                material_resources.append(resource)
        material_results = asyncio.run(Helper.fetch(material_resources, "material", 0, 5), debug=True)  
        
        results = logic_results + law_results + trumpet_results + blog_results + material_results

        slow_resources = []
        for result in results:
            if result.data == None:
                slow_resources.append(result.name)

        for resource in resources:
            if resource.name in slow_resources:
                resource.disabled = True
                resource.save()
                print(f'{resource.name} was disabled!')

        # update Federation.timestamp
        federation = Federation.objects.all()[0] ###
        federation.save()

    @classmethod
    def populate(self) -> None:
        """ Populate federation with all resources """

        self.populate_skosmosinstances()
        self.populate_sparqlendpoints()
        self.populate_lobidresources()
        self.populate_ldapiendpoints()

    @classmethod
    def populate_skosmosinstances(self) -> None: 
        """ Populate federation with Skosmos instances and their queries """

        SkosmosInstance.objects.all().delete()

        skosmosinstances = load_workbook(LOCAL_APP_PATH + "/fixtures/skosmosinstances.xlsx")
        for ws in skosmosinstances:
            
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):   
                instance = SkosmosInstance(federation=Federation.objects.all()[0],
                                           name=row[0],
                                           url=row[1],
                                           context=row[2])
                instance.save()

                query = SkosmosQuery(skosmosinstance=instance, # constructs query of cat 0
                                     querystring="/rest/v1/search?query=",
                                     description="NA",
                                     category=0,
                                     timeout=row[3])
                query.save()

    @classmethod
    def populate_sparqlendpoints(self) -> None:
        """ Populate federation with SPARQL endpoints and their queries """

        SparqlEndpoint.objects.all().delete()

        sparqlendpoints = load_workbook(LOCAL_APP_PATH + "/fixtures/sparqlendpoints.xlsx")
        for ws in sparqlendpoints:

            endpoint = SparqlEndpoint(federation=Federation.objects.all()[0],
                                      name=ws['A3'].value,
                                      url=ws['B3'].value,
                                      context=ws['C3'].value)
            endpoint.save()
            
            for row in ws.iter_rows(min_row=3, min_col=4, max_col=7, values_only=True):
                query = SparqlQuery(sparqlendpoint=endpoint,
                                    querystring=row[0],
                                    description=row[1],
                                    category=row[2],
                                    timeout=row[3])
                query.save()

    @classmethod
    def populate_lobidresources(self) -> None:
        """ Populate federation with Lobid resources and their queries """

        LobidResource.objects.all().delete()

        lobidresources = load_workbook(LOCAL_APP_PATH + "/fixtures/lobidresources.xlsx")
        for ws in lobidresources:

            for row in ws.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):   
                resource = LobidResource(federation=Federation.objects.all()[0],
                                           name=row[0],
                                           url=row[1],
                                           context=row[2])
                resource.save()

                query = LobidQuery(lobidresource=resource, # constructs query of cat 0
                                     querystring="/search?q=preferredName%3A!!!SEARCHWORD!!!+OR+variantName%3A!!!SEARCHWORD!!!&size=100&format=json",
                                     description="NA",
                                     category=0,
                                     timeout=row[3])
                query.save()

    @classmethod
    def prepare_ldapiendpoints(self) -> None:
        """ Prepare ldapiendpoints.xlsx """

        ldapiendpoints = load_workbook(LOCAL_APP_PATH + "/fixtures/ldapiendpoints.xlsx")

        for ws in ldapiendpoints:
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=False):
                rawname = row[1].value
                rawname = rawname.replace("http://vocabs.ands.org.au/repository/api/lda/", "")
                namelist = rawname.split("/")
                if namelist[len(namelist)-2] == "current":
                    row[0].value = namelist[len(namelist)-3]
                else:
                    row[0].value = namelist[len(namelist)-2]

        ldapiendpoints.save(LOCAL_APP_PATH + "/fixtures/ldapiendpoints.xlsx")

    @classmethod
    def populate_ldapiendpoints(self) -> None: 
        """ Populate federation with Linked-Data-API endpoints and their queries """

        self.prepare_ldapiendpoints()
        try:
            LdapiEndpoint.objects.all().delete()
        except:
            pass
        
        ldapiendpoints = load_workbook(LOCAL_APP_PATH + "/fixtures/ldapiendpoints.xlsx")
        for ws in ldapiendpoints:
            
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):   
                endpoint = LdapiEndpoint(federation=Federation.objects.all()[0],
                                         name=row[0],
                                         url=row[1],
                                         disabled=True,
                                         context=row[2])
                endpoint.save()

                query = LdapiQuery(ldapiendpoint=endpoint, # constructs query of cat 0
                                     querystring="?labelcontains=",
                                     description="NA",
                                     category=0,
                                     timeout=row[3])
                query.save()


                
            

        


        

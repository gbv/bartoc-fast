""" models_main.py """

import time # dev
import asyncio

from django.db import models # https://docs.djangoproject.com/en/2.2/ref/models/fields/#django.db.models.Field
from aiohttp import ClientSession, ClientTimeout
from openpyxl import load_workbook

from ..utility import Database, Entry, Result, LOCAL_APP_PATH

class Query(models.Model): # confusing, same name as schema.Query
    """ Abstract query class """
    
    querystring = models.CharField(max_length=100)
    description = models.CharField(max_length=100)
    category = models.IntegerField()
    timeout = models.IntegerField()

    class Meta:
        abstract = True

class Resource(models.Model):
    """ Abstract resource class """

    federation = models.ForeignKey("Federation", on_delete=models.CASCADE)
    name = models.CharField(max_length=100)
    url = models.URLField()
    disabled = models.BooleanField(default=False)
    context = models.CharField(max_length=200) #special, wildcard

    class Meta:
        abstract = True

    def __str__(self) -> str:
        return f'{self.name}'

    async def main(self,
                   session: ClientSession,
                   searchword: str,
                   category: int = 0) -> Result:
        """ Coroutine: send query to resource """

        try:
            result = await self.search(session, searchword, category)
            return result
        
        except asyncio.TimeoutError:
            print(f'FETCH {self.name} ran out of time!') # dev
            return Result(self.name, None, category)

class Federation(models.Model):
    """ The federation of resources (there is only one) """
    
    timestamp = models.DateTimeField(auto_now=True)

    class Meta:
        verbose_name_plural = 'federation'

    def __str__(self) -> str:
        return 'Main federation'

    def get_timestamp(self):
        return self.timestamp.strftime("%Y-%m-%d %H:%M:%S")

    def populate_skosmosinstances(self) -> None: 
        """ Populate federation with Skosmos instances and their queries """

        SkosmosInstance.objects.all().delete()

        skosmosinstances = load_workbook(LOCAL_APP_PATH + "/fixtures/skosmosinstances.xlsx")
        for ws in skosmosinstances:
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):   
                instance = SkosmosInstance(federation=self,
                                           name=row[0],
                                           url=row[1],
                                           context=row[2])
                instance.save()

                query = SkosmosQuery(skosmosinstance=instance, # constructs query of cat 0
                                     querystring="/rest/v1/search?query=",
                                     description="NA",
                                     category=0,
                                     timeout=row[3])
                query.save()

    def populate_sparqlendpoints(self) -> None:
        """ Populate federation with SPARQL endpoints and their queries """

        SparqlEndpoint.objects.all().delete()

        sparqlendpoints = load_workbook(LOCAL_APP_PATH + "/fixtures/sparqlendpoints.xlsx")
        for ws in sparqlendpoints:

            endpoint = SparqlEndpoint(federation=self,
                                      name=ws['A3'].value,
                                      url=ws['B3'].value,
                                      context=ws['C3'].value)
            endpoint.save()
            
            for row in ws.iter_rows(min_row=3, min_col=4, max_col=7, values_only=True):
                query = SparqlQuery(sparqlendpoint=endpoint,
                                    querystring=row[0],
                                    description=row[1],
                                    category=row[2],
                                    timeout=row[3])
                query.save()

    def populate(self) -> None:
        """ Populate federation with all resources """

        self.populate_skosmosinstances()
        self.populate_sparqlendpoints()

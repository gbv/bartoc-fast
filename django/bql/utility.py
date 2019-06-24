""" utility.py """

from __future__ import annotations # see https://www.python.org/dev/peps/pep-0563/
from typing import List, Set, Dict, Tuple, Optional

from openpyxl import load_workbook

class Database:
    """ A database """
    
    def __init__(self, entries: set) -> None:
        self.entries = entries

    def set_entries(self, entries: set) -> None:
        self.entries = entries

    def add_entries(self, entries: set) -> None:
        self.entries = entries.union(self.entries)

class Entry:
    """ A named and linked entry of a database """
    
    def __init__(self, name: str, url: str) -> None:
        self.name = name
        self.url = url

    def set_name(self, name: str) -> None:
        self.name = name
         
    def set_url(self, url: str) -> None:
        self.url = url

class LanguageDatabase(Database):
    """ A collection of languages """
    
    def __init__(self, entries: set = {}) -> None:
        Database.__init__(self, entries)
        self.setup() # setup is called on initializing class

    def setup(self) -> None:
        """ Populates the database with languages """
        
        wb = load_workbook("//itsc-pg2.storage.p.unibas.ch/ub-home$/hinder0000/Documents/GitHub/bartoc-graphql/django/bql/fixtures/languages.xlsx") # required for unittest; fixtures/languages.xlsx in production
        for ws in wb:
            vocabulary = []
            for row in ws.iter_rows(min_row=3, min_col=4, max_col=5, values_only=True):
                uri = row[0]
                name = row[1]
                vocabulary.append(Word(uri, name))
            self.add_entries({Language(ws['A3'].value, Namespace(ws['B3'].value, ws['C3'].value), vocabulary)})
        
class Language():
    """ A Language """
    
    def __init__(self, name: str = None, namespace: Namespace = None, vocabulary: List[Word] = None) -> None:
        self.name = name
        self.namespace = namespace
        self.vocabulary = vocabulary

    def flat(self, word):
        """ Return flattened word """
        
        try:
            assert word in self.vocabulary
        except AssertionError:
            return None
        return self.namespace.uri + word.name

class Namespace():
    """ A namespace """
    
    def __init__(self, uri: str = None, prefix: str = None) -> None:
        self.uri = uri
        self.prefix = prefix    
        
class Word():
    """ A word """

    def __init__(self, uri: str = None, name: str = None) -> None:
        self.uri = uri
        self.name = name


#test()
        
# iso: http://purl.org/iso25964/skos-thes#

""""
    class GlobalResult(graphene.ObjectType):
    # Result of a global query 
    prefLabel: [String] "http://www.w3.org/2004/02/skos/core#prefLabel"
    altLabel: [String]  "http://www.w3.org/2004/02/skos/core#altLabel"
    hiddenLabel: [String]  "http://www.w3.org/2004/02/skos/core#hiddenLabel"
    broader: [Concept] "http://www.w3.org/2004/02/skos/core#broader"
    narrower: [Concept] "http://www.w3.org/2004/02/skos/core#narrower"
    related: [Concept] "http://www.w3.org/2004/02/skos/core#related"
    broadMatch: [Concept] "http://www.w3.org/2004/02/skos/core#broadMatch"
    closeMatch: [Concept]  "http://www.w3.org/2004/02/skos/core#closeMatch"
    narrowMatch: [Concept] "http://www.w3.org/2004/02/skos/core#narrowMatch"
    relatedMatch: [Concept] "http://www.w3.org/2004/02/skos/core#relatedMatch"
    exactMatch: [Concept] "http://www.w3.org/2004/02/skos/core#exactMatch"
    note: [String] "http://www.w3.org/2004/02/skos/core#note"
    definition: [String] "http://www.w3.org/2004/02/skos/core#inScheme"
    inScheme: [Scheme] "http://www.w3.org/2004/02/skos/core#inScheme"
    topConceptOf: [Scheme]  "http://www.w3.org/2004/02/skos/core#topConceptOf"

    uri = graphene.String(description='URI')
    type_ = graphene.List(graphene.String, description='type')

    lang = graphene.String(description='language')
    vocab = graphene.String(description='vocabulary')
    exvocab = graphene.String(description='???')
    """

"""    def __init__(self, name: str = None, namespace: Namespace = None, vocabulary: List[Word] = None) -> None:
        self.name = name
        self.namespace = namespace
        self.vocabulary = vocabulary"""

